import os
import wikipedia
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import string

def description_lookup(search_string, req_type):
    if (req_type):
        return category_lookup(search_string)
    else:
        return company_lookup(search_string)

# scrape Wikipedia for company description
def company_lookup(company_search_string):

    google_search = "http://www.google.com/search?q='" + company_search_string + "'+Company+Wikipedia"
    headers = {'User-Agent': 'Mozilla/5.0'}
    response = requests.get(google_search, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    wiki_request = company_search_string
    for link in soup.find_all('a'):
        if " - Wiki" in link.get_text():
            wiki_request = link.get_text()[:link.get_text().find(" - Wiki")]
            if not (company_search_string.lower() in wiki_request.lower()):
                wiki_request = company_search_string
            break

    try:
        return wikipedia.summary(wiki_request, sentences = 1)
    except:
        return " "

# scrape Investopedia for category definition
def category_lookup(category_search_string):

    base_url = "http://www.investopedia.com/terms/"
    request_url = base_url + category_search_string[0] + "/" + category_search_string.replace(" ", "_") + ".asp"

    try:
        response = requests.get(request_url)
        soup = BeautifulSoup(response.content, "html.parser")
        # find div that contains the definition
        target_div = soup.find('div', class_='content-box content-box-term')
        paragraph = target_div.find('p').get_text()

        # if definition is too long, only display first 2 sentences
        sentences = paragraph.split('.')
        if len(sentences) > 1:
            return sentences[0]
        else:
            return paragraph
    except:
        # return blank if definition not found
        return " "

script_path = os.path.dirname('__file__')
project_dir = os.path.abspath(os.path.join(script_path,'..','..','..','FintechExplorerProject'))
sys.path.insert(0, project_dir)
os.environ['DJANGO_SETTINGS_MODULE'] = 'FintechExplorerProject.settings'

wb = load_workbook(filename = 'static\Manage\Taxonomy.xlsx', data_only=True)
worksheet = wb.get_sheet_by_name(name = 'data')

printable = set(string.printable)
Loaded_Taxonomy = []
index = 2
next_row_exists = True
while (next_row_exists):
    try:
        if ((index % 20) == 0):
            print ("Reading line #" + str(index))

        id_imp = int(worksheet.cell(row=index, column=1).value)
        parent_imp = int(worksheet.cell(row=index, column=2).value)
        name_imp = worksheet.cell(row=index, column=3).value
        isCategory_imp = bool(worksheet.cell(row=index, column=4).value)
        description_imp = worksheet.cell(row=index, column=5).value

        if (description_imp is None):
            description_imp = description_lookup(str(name_imp),bool(isCategory_imp)).replace('"',"'").replace('’',"'")
            filter(lambda x: x in printable, description_imp)

        Loaded_Taxonomy.append([id_imp,parent_imp,name_imp,isCategory_imp,description_imp])
        #print ([id_imp,parent_imp,name_imp,isCategory_imp,description_imp])

        next_row_exists = False
        index = index + 1
        if (worksheet.cell(row=index, column=1).value and worksheet.cell(row=index, column=2).value and worksheet.cell(row=index, column=3).value):
            next_row_exists = True

    except Exception as e:
        raise type(e)('An error happened while importing from excel file. Check that file is correct.')

from FintechExplorerApp.models import Taxonomy_Node
for Node in Loaded_Taxonomy:
    DB_Node = Taxonomy_Node(Identifier=Node[0], Parent=Node[1], Name=Node[2], IsCategory=Node[3], Description=Node[4])
    DB_Node.save()

print ("All done")